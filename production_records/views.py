# production_records/views.py
from django.http import HttpResponse
import openpyxl
from .models import ProductionRecord, Expense, WindowFactoryRecord, WindowExpense
from .forms import (ProductionRecordForm, ExpenseForm, WindowFactoryRecordForm, LoginForm, WindowExpenseForm,
                    UserProfileForm, UserForm, ProfileForm, Profile)
from django.urls import reverse_lazy
from django.views.generic.edit import FormView
from django.db.models import Q, Sum
from django.shortcuts import get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required


def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('index')
    else:
        form = AuthenticationForm()
    return render(request, 'accounts/login.html', {'form': form})


@login_required
def calculator(request):
    return render(request, 'calc.html')


@login_required
def home(request):
    return render(request, 'home.html')


@login_required
def profile(request):
    # Проверка наличия профиля у пользователя
    if not hasattr(request.user, 'profile'):
        # Если профиля нет, создаем его
        profile = Profile.objects.create(user=request.user)
    else:
        profile = request.user.profile

    if request.method == 'POST':
        # Обработка формы для загрузки фотографии
        picture_form = UserProfileForm(request.POST, request.FILES, instance=profile)
        if picture_form.is_valid():
            picture_form.save()

        # Обработка формы для редактирования данных пользователя
        user_form = UserForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return redirect('profile')

    else:
        # Если запрос GET, создаем экземпляры форм для отображения
        picture_form = UserProfileForm(instance=profile)
        user_form = UserForm(instance=request.user)
        profile_form = ProfileForm(instance=profile)

    return render(request, 'profile.html', {
        'picture_form': picture_form,
        'user_form': user_form,
        'profile_form': profile_form,
    })


def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('record_list')
            else:
                form.add_error(None, 'Неверные учетные данные')
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})


def record_list(request):
    name_filter = request.GET.get('name_filter', '').lower()
    model_filter = request.GET.get('model_filter', '').lower()
    date_filter_start = request.GET.get('date_filter_start', '')
    date_filter_end = request.GET.get('date_filter_end', '')

    records = ProductionRecord.objects.all()

    if name_filter:
        records = records.filter(Q(name__icontains=name_filter))

    if model_filter:
        records = records.filter(Q(model__icontains=model_filter))

    if date_filter_start and date_filter_end:
        records = records.filter(date__range=[date_filter_start, date_filter_end])

    total_total = records.aggregate(total=Sum('total'))['total']

    return render(request, 'record_list.html', {
        'records': records,
        'name_filter': name_filter,
        'model_filter': model_filter,
        'date_filter_start': date_filter_start,
        'date_filter_end': date_filter_end,
        'total_total': total_total,
    })


class EditRecordView(FormView):
    template_name = 'edit_record.html'
    form_class = ProductionRecordForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = ProductionRecord.objects.get(pk=self.kwargs['pk'])
        return kwargs

    def form_valid(self, form):
        form.save()
        return redirect('record_list')

    def get_success_url(self):
        return reverse_lazy('record_list')


def edit_record(request, pk):
    record = get_object_or_404(ProductionRecord, pk=pk)

    if request.method == 'POST':
        form = ProductionRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('record_list')
    else:
        form = ProductionRecordForm(instance=record)

    return render(request, 'edit_record.html', {'form': form})


def add_record(request):
    if request.method == 'POST':
        form = ProductionRecordForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('record_list')
    else:
        form = ProductionRecordForm()
    return render(request, 'add_record.html', {'form': form})


def export_records(request):
    name_filter = request.GET.get('name_filter', '')
    model_filter = request.GET.get('model_filter', '')
    date_filter_start = request.GET.get('date_filter_start', '')
    date_filter_end = request.GET.get('date_filter_end', '')

    # Фильтрация записей
    records = ProductionRecord.objects.all()

    if name_filter:
        records = records.filter(name__iexact=name_filter)

    if model_filter:
        records = records.filter(model__iexact=model_filter)

    if date_filter_start and date_filter_end:
        records = records.filter(date__range=[date_filter_start, date_filter_end])

    # Экспорт в Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (f'attachment; filename="production_records_{name_filter}_{model_filter}_'
                                       f'{date_filter_start}_{date_filter_end}.xlsx"')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Заголовки
    headers = ["Дата", "Модель", "Мастер", "Кол-во", "Принял", "Цена", "Итого"]
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet[f"{col_letter}1"] = header

    # Данные
    for row_num, record in enumerate(records, 2):
        worksheet[f"A{row_num}"] = record.date.strftime("%Y-%m-%d")
        worksheet[f"B{row_num}"] = record.model
        worksheet[f"C{row_num}"] = record.name
        worksheet[f"D{row_num}"] = record.quantity
        worksheet[f"E{row_num}"] = record.received_by
        worksheet[f"F{row_num}"] = record.price
        worksheet[f"G{row_num}"] = record.total

    workbook.save(response)
    return response


def window_factory_record_list(request):
    name_filter = request.GET.get('name_filter', '').lower()
    model_filter = request.GET.get('model_filter', '').lower()
    date_filter_start = request.GET.get('date_filter_start', '')
    date_filter_end = request.GET.get('date_filter_end', '')

    windows = WindowFactoryRecord.objects.all()  # Исправлено здесь

    if name_filter:
        windows = windows.filter(Q(name__icontains=name_filter))

    if model_filter:
        windows = windows.filter(Q(model__icontains=model_filter))

    if date_filter_start and date_filter_end:
        windows = windows.filter(date__range=[date_filter_start, date_filter_end])

    total_total = windows.aggregate(total=Sum('total'))['total']

    return render(request, 'window_list.html', {
        'windows': windows,  # Исправлено здесь
        'name_filter': name_filter,
        'model_filter': model_filter,
        'date_filter_start': date_filter_start,
        'date_filter_end': date_filter_end,
        'total_total': total_total,
    })


class EditWindowFactoryRecordView(FormView):
    template_name = 'edit_window.html'
    form_class = WindowFactoryRecordForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = WindowFactoryRecord.objects.get(pk=self.kwargs['pk'])
        return kwargs

    def form_valid(self, form):
        form.save()
        return redirect('window_factory_record_list')

    def get_success_url(self):
        return reverse_lazy('window_factory_record_list')


def edit_window_factory_record(request, pk):
    record = get_object_or_404(WindowFactoryRecord, pk=pk)

    if request.method == 'POST':
        form = WindowFactoryRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('window_factory_record_list')
    else:
        form = WindowFactoryRecordForm(instance=record)

    return render(request, 'edit_window.html', {'form': form})


def add_window_factory_record(request):
    if request.method == 'POST':
        form = WindowFactoryRecordForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('window_factory_record_list')
    else:
        form = WindowFactoryRecordForm()
    return render(request, 'add_window.html', {'form': form})


def export_windows(request):
    name_filter = request.GET.get('name_filter', '')
    model_filter = request.GET.get('model_filter', '')
    date_filter_start = request.GET.get('date_filter_start', '')
    date_filter_end = request.GET.get('date_filter_end', '')

    # Фильтрация записей
    windows = WindowFactoryRecord.objects.all()

    if name_filter:
        windows = windows.filter(name__iexact=name_filter)

    if model_filter:
        windows = windows.filter(model__iexact=model_filter)

    if date_filter_start and date_filter_end:
        windows = windows.filter(date__range=[date_filter_start, date_filter_end])

    # Экспорт в Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (f'attachment; filename="window_factory_records_{name_filter}_{model_filter}_'
                                       f'{date_filter_start}_{date_filter_end}.xlsx"')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Заголовки
    headers = ["Дата", "Модель", "Мастер", "Кол-во", "Принял", "Цена", "Итого"]
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet[f"{col_letter}1"] = header

    # Данные
    for row_num, window in enumerate(windows, 2):
        worksheet[f"A{row_num}"] = window.date.strftime("%Y-%m-%d")
        worksheet[f"B{row_num}"] = window.model
        worksheet[f"C{row_num}"] = window.name
        worksheet[f"D{row_num}"] = window.quantity
        worksheet[f"E{row_num}"] = window.received_by
        worksheet[f"F{row_num}"] = window.price
        worksheet[f"G{row_num}"] = window.total

    workbook.save(response)
    return response


# ---- EXPENSES ----
def expenses_list(request):
    date_filter_start = request.GET.get('date_filter_start', '')
    date_filter_end = request.GET.get('date_filter_end', '')
    model_filter = request.GET.get('model_filter', '').lower()
    fabric_filter = request.GET.get('fabric_filter', '').lower()

    expenses = Expense.objects.all()

    if date_filter_start and date_filter_end:
        expenses = expenses.filter(date__range=[date_filter_start, date_filter_end])

    if model_filter:
        expenses = expenses.filter(Q(model_name__icontains=model_filter))

    if fabric_filter:
        expenses = expenses.filter(Q(name_fabric__icontains=fabric_filter))

    total_total = expenses.aggregate(total=Sum('total'))['total']

    return render(request, 'expenses_list.html', {
        'expenses': expenses,
        'date_filter_start': date_filter_start,
        'date_filter_end': date_filter_end,
        'model_filter': model_filter,
        'fabric_filter': fabric_filter,
        'total_total': total_total,
    })


class EditExpenseView(FormView):
    template_name = 'edit_expense.html'
    form_class = ExpenseForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = Expense.objects.get(pk=self.kwargs['pk'])
        return kwargs

    def form_valid(self, form):
        form.save()
        return redirect('expenses_list')

    def get_success_url(self):
        return reverse_lazy('expenses_list')


def edit_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            return redirect('expenses_list')
    else:
        form = ExpenseForm(instance=expense)

    return render(request, 'edit_expense.html', {'form': form})


def add_expense(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('expenses_list')
    else:
        form = ExpenseForm()
    return render(request, 'add_expense.html', {'form': form})


def export_expenses(request):
    date_filter_start = request.GET.get('date_filter_start', '')
    date_filter_end = request.GET.get('date_filter_end', '')
    model_filter = request.GET.get('model_filter', '')
    fabric_filter = request.GET.get('fabric_filter', '')

    # Фильтрация расходов
    expenses = Expense.objects.all()

    if date_filter_start and date_filter_end:
        expenses = expenses.filter(date__range=[date_filter_start, date_filter_end])

    if model_filter:
        expenses = expenses.filter(model_name__iexact=model_filter)

    if fabric_filter:
        expenses = expenses.filter(name_fabric__iexact=fabric_filter)

    # Экспорт в Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (f'attachment; filename="expenses_{date_filter_start}_{date_filter_end}_{model_filter}_{fabric_filter}.xlsx"')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Заголовки
    headers = ["Дата", "Модель", 'Назв. ткани', "Ткань", "Фурнитура", "Нитки", "Другое", "Пошив", "Итого"]
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet[f"{col_letter}1"] = header

    # Данные
    for row_num, expense in enumerate(expenses, 2):
        worksheet[f"A{row_num}"] = expense.date.strftime("%d-%m-%Y")
        worksheet[f"B{row_num}"] = expense.model_name
        worksheet[f"C{row_num}"] = expense.name_fabric
        worksheet[f"D{row_num}"] = expense.fabric
        worksheet[f"E{row_num}"] = expense.accessories
        worksheet[f"F{row_num}"] = expense.threads
        worksheet[f"G{row_num}"] = expense.other
        worksheet[f"H{row_num}"] = expense.sewing
        worksheet[f"I{row_num}"] = expense.total

    try:
        workbook.save(response)
    except Exception as e:
        print(f"An error occurred: {e}")
    return response


def window_expenses_list(request):
    date_filter_start = request.GET.get('date_filter_start', '')
    date_filter_end = request.GET.get('date_filter_end', '')
    window_model_filter = request.GET.get('window_model_filter', '').lower()

    expenses = WindowExpense.objects.all()

    if date_filter_start and date_filter_end:
        expenses = expenses.filter(date__range=[date_filter_start, date_filter_end])

    if window_model_filter:
        expenses = expenses.filter(Q(window_model_name__icontains=window_model_filter))

    total_total = expenses.aggregate(total=Sum('total'))['total']

    return render(request, 'window_expenses_list.html', {
        'expenses': expenses,
        'date_filter_start': date_filter_start,
        'date_filter_end': date_filter_end,
        'window_model_filter': window_model_filter,
        'total_total': total_total,
    })


class EditWindowExpenseView(FormView):
    template_name = 'edit_window_expense.html'
    form_class = WindowExpenseForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = WindowExpense.objects.get(pk=self.kwargs['pk'])
        return kwargs

    def form_valid(self, form):
        form.save()
        return redirect('window_expenses_list')

    def get_success_url(self):
        return reverse_lazy('window_expenses_list')


def edit_window_expense(request, pk):
    expense = get_object_or_404(WindowExpense, pk=pk)

    if request.method == 'POST':
        form = WindowExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            return redirect('window_expenses_list')
    else:
        form = WindowExpenseForm(instance=expense)

    return render(request, 'edit_window_expense.html', {'form': form})


def add_window_expense(request):
    if request.method == 'POST':
        form = WindowExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('window_expenses_list')
    else:
        form = WindowExpenseForm()

    return render(request, 'add_window_expense.html', {'form': form})


def export_window_expenses(request):
    window_model_filter = request.GET.get('window_model_filter', '')
    date_filter_start = request.GET.get('date_filter_start', '')
    date_filter_end = request.GET.get('date_filter_end', '')

    expenses = WindowExpense.objects.all()

    if date_filter_start and date_filter_end:
        expenses = expenses.filter(date__range=[date_filter_start, date_filter_end])

    if window_model_filter:
        expenses = expenses.filter(window_model_name__iexact=window_model_filter)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (f'attachment; filename="window_expenses_{date_filter_start}_{date_filter_end}.xlsx"')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    headers = ["Дата", "Модель окна", "Стекло", "Рама", "Фурнитура", "Установка", "Другое", "Итого"]
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet[f"{col_letter}1"] = header

    for row_num, expense in enumerate(expenses, 2):
        worksheet[f"A{row_num}"] = expense.date.strftime("%Y-%m-%d")
        worksheet[f"B{row_num}"] = expense.window_model_name
        worksheet[f"C{row_num}"] = expense.glass
        worksheet[f"D{row_num}"] = expense.frame
        worksheet[f"E{row_num}"] = expense.fittings
        worksheet[f"F{row_num}"] = expense.installation
        worksheet[f"G{row_num}"] = expense.other
        worksheet[f"H{row_num}"] = expense.total

    workbook.save(response)
    return response
